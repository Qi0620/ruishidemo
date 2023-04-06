import os

import allure
import jsonpath
import openpyxl
import pytest

from ruishidemomo.excel_red.red_excel import RedExcel
from ruishidemomo.api_keyword.api_key import ApiKey
from ruishidemomo.fengzhuang.check_fengzhuang import CheckSql


class TestInterfaceExcel:
    models = RedExcel().read_excle()
    sqldict = {}

    @classmethod
    def setup_class(cls):
        global dfdict,h5dict,sheet,wb
        dfdict = {}
        h5dict = {}
        # wb = openpyxl.load_workbook("../data/test_ruis.xlsx")
        wb = openpyxl.load_workbook("./ruishidemomo/data/test_ruis.xlsx")
        sheet = wb["info"]

    @pytest.mark.parametrize("model", models)
    def test_sendhttp(self, model, **kwargs):
        print(model.idd,"modemodemode")
        if model.desc is not None:
            # 动态获取用例标题
            allure.dynamic.title(model.desc)
        if model.story is not None:
            # 动态获取story模块名
            allure.dynamic.story(model.story)
        if model.feature is not None:
            # 动态获取大模块
            allure.dynamic.feature(model.feature)
        if model.backup is not None:
            # 动态获取备注
            allure.dynamic.description(model.backup)
        if model.level is not None:
            allure.dynamic.severity(model.level)

        ak = ApiKey()

        if model.headers:
            model.headers = eval(model.headers)

        if model.data:
            model.data = model.data.replace('null', 'None')
            model.data = eval(model.data)
        print(model.data, "============")

        # 第三行接口 商品上传图片中 请求参数需要物料编码   去数据库查询
        # print(self.models.index(model), "----------------")
        if model.desc == "上传商品图片":
            # if self.models.index(model) == 1:
            ck = CheckSql(host="fat-bj-china-b2c-02.chfa9nngsipy.rds.cn-north-1.amazonaws.com.cn",
                          user="dml_user", password="tiens_123", port=3306, database="ruishi_goodscenter")
            sql = ck.chSql("select id from t_product where name='金积分EBSLK是0' ORDER BY creation_time desc", 1)
            # print(sql)
            # print(type(sql))
            newsql = sql[0][0]
            print(newsql)
            # 将查询到的物料编码放到一个字典中
            self.sqldict["productId"] = newsql
            # 循环这个字典添加到data中
            for i in self.sqldict:
                print(i)
                model.data[i] = self.sqldict[i]
            print(model.data)

        # 小程序登陆请求参数thirdtype值是1   eval无法转换，单独去添加一下
        # if model.desc == "小程序登陆":
        #     thirdty = dfdict["thirdType"]
        #     model.data["thirdType"] = thirdty
        # print(model.data, '=======================')

        # 反射+解包
        res = ak.sendhttp(url=model.url, method=model.method, data=model.data,
                          headers=model.headers, par_type=model.par_type)
        print(res.json())

        # 数据抽取
        if model.extract is not None:
            exval = model.extract.split(";")
            ex = model.is_need.split(";")
            for i in range(len(exval)):
                dfdict[exval[i]] = str(ak.get_data(res, ex[i]))
        print(dfdict,"数据抽取++++++++++++")

        if model.desc == "金积分充值查询积分余额":
            print(type(dfdict['balance']))
            ck = CheckSql(host="fat-bj-china-b2c-02.chfa9nngsipy.rds.cn-north-1.amazonaws.com.cn",
                          user="dml_user", password="tiens_123", port=3306, database="ruishi_user_center")
            if dfdict['balance'] == "0.0":
                sql = f"update ruishi_user_center.user_balance set balance=10000.0000,frozen=0.0000," \
                      f"useful=10000.0000 where user_id={dfdict['uid']}"
                ck.AddUpDel(sql)
                print("充值完成")
            else:
                sql = f"select balance from ruishi_user_center.user_balance where user_id={dfdict['uid']}"
                money = ck.chSql(sql, 1)
                print("还有钱，当前金积分余额为：{}".format(money[0][0]))

        if model.desc == "银积分充值查询积分余额":
            ck = CheckSql(host="fat-bj-china-b2c-02.chfa9nngsipy.rds.cn-north-1.amazonaws.com.cn",
                          user="dml_user", password="tiens_123", port=3306, database="ruishi_user_center")
            if dfdict['pointsUseful'] == "0.00":
                sql = f"update ruishi_user_center.user_points set compound_points_total=10000.0000," \
                      f"compound_points_remainder=10000.0000,compound_points_useful=10000.0000 where user_id={dfdict['uid']}"
                ck.AddUpDel(sql)
                print("充值完成")
            else:
                sql = f"select compound_points_useful from ruishi_user_center.user_points where user_id={dfdict['uid']}"
                money = ck.chSql(sql, 1)
                print("还有钱，当前银积分余额为：{}".format(money[0][0]))


        # ,"captcha":"587364"
        if model.desc == "H5发送验证码":
            ck = CheckSql(host="fat-bj-china-b2c-02.chfa9nngsipy.rds.cn-north-1.amazonaws.com.cn",
                          user="dml_user", password="tiens_123", port=3306, database="ruishi_core_service")
            sql = "select content from ruishi_core_service.base_sms_record where mobile = '18833333333' ORDER BY created_time desc limit 1"
            captcha = ck.chSql(sql, 1)[0][0]
            print(captcha)  #字符串格式   进行切片取到验证码
            dfdict["captcha"] = captcha[21:27]
        print(dfdict,"=============================================")
        print(model.headers, "************************************")

        # 反显
        with allure.step("断言反显"):
            r = model.idd + 1    # 加一个编号   通过编号定位到当前执行行
            try:
                if model.assert_values == "true":
                    model.assert_values = model.assert_values.replace("true", "True")
                if model.assert_options == "等于":
                    print(model.assert_values ,jsonpath.jsonpath(res.json(), '$..' + model.assert_data)[0],"aaaaaaaaaaaaaaaaaaaa")
                    if model.assert_values == str(jsonpath.jsonpath(res.json(), '$..' + model.assert_data)[0]):
                        print(model.idd,"iddiddiddiddiddiddidd")
                        print("执行通过")
                        sheet.cell(r, 18).value = '通过'
                    else:
                        print("执行失败")
                        sheet.cell(r, 18).value = '失败'
                if model.assert_options == "包含":
                    print(type(res.json()))
                    if model.assert_data in res.json():
                        print("success在字典中")
                        print(model.idd, "bbbbbbbbbbbbbbb")
                        print("执行通过")
                        sheet.cell(r, 18).value = '通过'
                    else:
                        print("执行失败")
                        sheet.cell(r, 18).value = '失败'
                if model.assert_options == "大于":
                    print(type(res.json()))
                    print(model.assert_values, jsonpath.jsonpath(res.json(), '$..' + model.assert_data)[0],
                          "cccccccccccccccccc")
                    print(type(model.assert_values),type(jsonpath.jsonpath(res.json(), '$..' + model.assert_data)[0]))
                    if model.assert_values < jsonpath.jsonpath(res.json(), '$..' + model.assert_data)[0]:
                        print(model.idd, "cccccccccccccccccc")
                        print("执行通过")
                        sheet.cell(r, 18).value = '通过'
                    else:
                        print("执行失败")
                        sheet.cell(r, 18).value = '失败'
                if model.assert_options == "小于":
                    print(type(res.json()))
                    print(model.assert_values, jsonpath.jsonpath(res.json(), '$..' + model.assert_data)[0],
                          "ddddddddddddddd")
                    print(type(model.assert_values),type(jsonpath.jsonpath(res.json(), '$..' + model.assert_data)[0]))
                    if jsonpath.jsonpath(res.json(), '$..' + model.assert_data)[0] == "0":
                        print(model.idd, "ddddddddddddddd")
                        print("执行通过")
                        sheet.cell(r, 18).value = '通过'
                    else:
                        print("执行失败")
                        sheet.cell(r, 18).value = '失败'
            except Exception as e:
                sheet.cell(r, 18).value = e
            finally:
                wb.save("./ruishidemomo/data/test_ruis.xlsx")


if __name__ == '__main__':
    pytest.main(['-v', '--alluredir', './result', '--clean-alluredir'])
    os.system('allure generate ./result/ -o ./allure_report --clean')